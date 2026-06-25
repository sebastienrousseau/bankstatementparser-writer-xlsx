# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Tests for the bankstatementparser-writer-xlsx writer."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from bankstatementparser import Transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from bankstatementparser_writer_xlsx import __version__, write_xlsx
from bankstatementparser_writer_xlsx.writer import _MAX_COLUMN_WIDTH


def _transactions() -> list[Transaction]:
    """Build two Transaction fixtures exercising Decimal/date fields."""
    return [
        Transaction(
            account_id="DE89370400440532013000",
            currency="EUR",
            amount=Decimal("1250.75"),
            booking_date=date(2026, 6, 1),
            value_date=date(2026, 6, 2),
            description="ACME GmbH invoice 123",
            reference="REF-1",
            transaction_id="TX-1",
            counterparty="ACME GmbH",
            source="statement.csv",
        ),
        Transaction(
            account_id="DE89370400440532013000",
            currency="EUR",
            amount=Decimal("-42.00"),
            booking_date=date(2026, 6, 3),
            description="Coffee Shop",
            counterparty="Coffee Shop",
            source="statement.csv",
        ),
    ]


def test_version_exposed() -> None:
    """The package exposes a non-empty semantic-style version string."""
    assert isinstance(__version__, str)
    assert __version__ == "0.0.12"


def test_returns_path_and_writes_file(tmp_path: Path) -> None:
    """The function returns the output path and creates a non-empty file."""
    out = tmp_path / "out.xlsx"

    result = write_xlsx(_transactions(), out)

    assert result == out
    assert isinstance(result, Path)
    assert out.exists()
    assert out.stat().st_size > 0


def test_transaction_input_header_and_order(tmp_path: Path) -> None:
    """list[Transaction] yields the stable header in field order."""
    out = tmp_path / "tx.xlsx"

    write_xlsx(_transactions(), out)

    sheet = load_workbook(out)["Transactions"]
    header = [c.value for c in sheet[1]]
    # First columns follow the stable Transaction order.
    assert header[:6] == [
        "account_id",
        "currency",
        "amount",
        "booking_date",
        "value_date",
        "description",
    ]
    assert "transaction_hash" in header


def test_transaction_decimal_and_date_coercion(tmp_path: Path) -> None:
    """Decimal becomes float; date round-trips as a native date cell."""
    out = tmp_path / "coerce.xlsx"

    write_xlsx(_transactions(), out)

    sheet = load_workbook(out)["Transactions"]
    header = [c.value for c in sheet[1]]
    amount_col = header.index("amount")
    booking_col = header.index("booking_date")
    first_data = [c.value for c in sheet[2]]

    amount = first_data[amount_col]
    assert isinstance(amount, float)
    assert amount == pytest.approx(1250.75)

    booking = first_data[booking_col]
    assert isinstance(booking, datetime)
    assert booking.date() == date(2026, 6, 1)


def test_header_row_is_bold(tmp_path: Path) -> None:
    """The header row is rendered bold."""
    out = tmp_path / "bold.xlsx"

    write_xlsx(_transactions(), out)

    sheet = load_workbook(out)["Transactions"]
    assert all(cell.font.bold for cell in sheet[1])


def test_dataframe_input_preserves_columns(tmp_path: Path) -> None:
    """A DataFrame keeps its own column order and row values."""
    frame = pd.DataFrame(
        {
            "date": [date(2026, 6, 1), date(2026, 6, 2)],
            "description": ["Salary", "Rent"],
            "amount": [Decimal("3000.00"), Decimal("-1200.00")],
        }
    )
    out = tmp_path / "df.xlsx"

    write_xlsx(frame, out)

    sheet = load_workbook(out)["Transactions"]
    header = [c.value for c in sheet[1]]
    assert header == ["date", "description", "amount"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    assert rows[0][1] == "Salary"
    assert rows[0][2] == pytest.approx(3000.0)


def test_dict_input_union_of_keys_first_seen_order(tmp_path: Path) -> None:
    """Dict records use the union of keys in first-seen order."""
    records = [
        {"date": "2026-06-01", "amount": 10},
        {"date": "2026-06-02", "amount": 20, "memo": "extra"},
    ]
    out = tmp_path / "dicts.xlsx"

    write_xlsx(records, out)

    sheet = load_workbook(out)["Transactions"]
    header = [c.value for c in sheet[1]]
    assert header == ["date", "amount", "memo"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # The first record has no "memo" → blank cell.
    assert rows[0] == ("2026-06-01", 10, None)
    assert rows[1] == ("2026-06-02", 20, "extra")


def test_custom_sheet_name(tmp_path: Path) -> None:
    """The sheet_name argument titles the transactions sheet."""
    out = tmp_path / "named.xlsx"

    write_xlsx(_transactions(), out, sheet_name="Ledger")

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["Ledger"]


def test_summary_sheet(tmp_path: Path) -> None:
    """A summary mapping is written to a second Summary sheet."""
    summary = {
        "account_id": "DE89370400440532013000",
        "transaction_count": 2,
        "total_amount": Decimal("1208.75"),
        "currency": "EUR",
    }
    out = tmp_path / "summary.xlsx"

    write_xlsx(_transactions(), out, summary=summary)

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["Transactions", "Summary"]
    sheet = workbook["Summary"]
    header = [c.value for c in sheet[1]]
    assert header == ["Key", "Value"]
    assert all(cell.font.bold for cell in sheet[1])
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("account_id", "DE89370400440532013000")
    assert rows[1] == ("transaction_count", 2)
    # Decimal total coerced to float.
    assert rows[2][0] == "total_amount"
    assert rows[2][1] == pytest.approx(1208.75)


def test_empty_list_writes_empty_workbook(tmp_path: Path) -> None:
    """An empty list writes a valid workbook with no header and no rows."""
    out = tmp_path / "empty.xlsx"

    result = write_xlsx([], out)

    assert result == out
    sheet = load_workbook(out)["Transactions"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows == []


def test_empty_dataframe_writes_header_only(tmp_path: Path) -> None:
    """An empty DataFrame with columns writes a header-only sheet."""
    frame = pd.DataFrame(columns=["date", "amount"])
    out = tmp_path / "empty-df.xlsx"

    write_xlsx(frame, out)

    sheet = load_workbook(out)["Transactions"]
    header = [c.value for c in sheet[1]]
    assert header == ["date", "amount"]
    data = list(sheet.iter_rows(min_row=2, values_only=True))
    assert data == []


def test_accepts_string_path(tmp_path: Path) -> None:
    """The path argument accepts a str as well as a Path."""
    out = tmp_path / "string.xlsx"

    result = write_xlsx(_transactions(), str(out))

    assert result == out


def test_non_amount_value_stringified(tmp_path: Path) -> None:
    """An exotic value type falls back to its string form."""
    out = tmp_path / "exotic.xlsx"

    write_xlsx([{"value": {"nested": 1}, "flag": True}], out)

    sheet = load_workbook(out)["Transactions"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0][0] == str({"nested": 1})
    # Booleans pass through unchanged (not coerced to int/float).
    assert rows[0][1] is True


def test_unsupported_top_level_type_raises_typeerror(tmp_path: Path) -> None:
    """A non-sequence, non-DataFrame argument raises TypeError."""
    with pytest.raises(TypeError):
        write_xlsx("not-a-table", tmp_path / "bad.xlsx")


def test_unsupported_item_type_raises_valueerror(tmp_path: Path) -> None:
    """A sequence item that is neither dict nor Transaction raises."""
    with pytest.raises(ValueError):
        write_xlsx([42], tmp_path / "bad-item.xlsx")


# ----------------------------------------------------------------------
# Edge cases — read the workbook back and assert exact cells.
# ----------------------------------------------------------------------


def test_empty_list_has_no_header_or_rows(tmp_path: Path) -> None:
    """An empty list writes a workbook with neither header nor data."""
    out = tmp_path / "empty-list.xlsx"

    write_xlsx([], out)

    sheet = load_workbook(out)["Transactions"]
    assert list(sheet.iter_rows(values_only=True)) == []
    # No bold header cell was written at all.
    assert sheet.max_row == 1
    assert sheet["A1"].value is None


def test_empty_dataframe_with_columns_writes_header_only(
    tmp_path: Path,
) -> None:
    """An empty DataFrame carrying columns writes a header-only sheet."""
    frame = pd.DataFrame(columns=["account_id", "amount", "currency"])
    out = tmp_path / "empty-df-cols.xlsx"

    write_xlsx(frame, out)

    sheet = load_workbook(out)["Transactions"]
    header = [cell.value for cell in sheet[1]]
    assert header == ["account_id", "amount", "currency"]
    assert all(cell.font.bold for cell in sheet[1])
    assert list(sheet.iter_rows(min_row=2, values_only=True)) == []


def test_unicode_and_overlong_string_cell(tmp_path: Path) -> None:
    """Unicode round-trips and an overlong string caps the column width."""
    long_text = "A" * 200
    out = tmp_path / "unicode.xlsx"

    write_xlsx([{"note": "Café Münster 日本語 €", "long": long_text}], out)

    sheet = load_workbook(out)["Transactions"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # Unicode content is preserved byte-for-byte on round-trip.
    assert rows[0][0] == "Café Münster 日本語 €"
    # The overlong string is stored verbatim (no truncation of content)...
    assert rows[0][1] == long_text
    assert len(rows[0][1]) == 200
    # ...but its column width is clamped to the 60-char cap.
    width = sheet.column_dimensions[get_column_letter(2)].width
    assert width == _MAX_COLUMN_WIDTH


def test_none_and_nan_cells_round_trip_as_blank(tmp_path: Path) -> None:
    """``None`` and ``NaN`` cells both read back as empty cells."""
    frame = pd.DataFrame({"amount": [1.0, None], "description": ["x", None]})
    dict_out = tmp_path / "nan-dict.xlsx"
    df_out = tmp_path / "nan-df.xlsx"

    write_xlsx([{"amount": float("nan"), "description": None}], dict_out)
    write_xlsx(frame, df_out)

    dict_rows = list(
        load_workbook(dict_out)["Transactions"].iter_rows(
            min_row=2, values_only=True
        )
    )
    # Explicit None and float NaN both serialise to an empty cell.
    assert dict_rows == [(None, None)]

    df_rows = list(
        load_workbook(df_out)["Transactions"].iter_rows(
            min_row=2, values_only=True
        )
    )
    assert df_rows[0] == (1.0, "x")
    assert df_rows[1] == (None, None)


def test_dict_records_with_missing_keys_union_first_seen(
    tmp_path: Path,
) -> None:
    """list[dict] with differing keys yields a first-seen-order union."""
    records = [
        {"a": 1, "b": 2},
        {"b": 20, "c": 30},
        {"a": 100, "d": 400},
    ]
    out = tmp_path / "union.xlsx"

    write_xlsx(records, out)

    sheet = load_workbook(out)["Transactions"]
    header = [cell.value for cell in sheet[1]]
    # Union of keys, in the order each key is first encountered.
    assert header == ["a", "b", "c", "d"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == (1, 2, None, None)
    assert rows[1] == (None, 20, 30, None)
    assert rows[2] == (100, None, None, 400)


def test_large_and_negative_decimal_amounts(tmp_path: Path) -> None:
    """Large and negative Decimals coerce to exact float values."""
    out = tmp_path / "decimals.xlsx"

    write_xlsx(
        [
            {"big": Decimal("123456789012.34")},
            {"big": Decimal("-9999.99")},
        ],
        out,
    )

    sheet = load_workbook(out)["Transactions"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert isinstance(rows[0][0], float)
    assert rows[0][0] == pytest.approx(123456789012.34)
    assert isinstance(rows[1][0], float)
    assert rows[1][0] == pytest.approx(-9999.99)


def test_date_and_datetime_cells(tmp_path: Path) -> None:
    """A ``date`` cell loses no day; a ``datetime`` keeps its time."""
    out = tmp_path / "datetimes.xlsx"

    write_xlsx(
        [{"d": date(2026, 6, 1), "dt": datetime(2026, 6, 1, 13, 30, 15)}],
        out,
    )

    sheet = load_workbook(out)["Transactions"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    d_cell, dt_cell = rows[0]
    # openpyxl reads a plain date back as a midnight datetime.
    assert isinstance(d_cell, datetime)
    assert d_cell == datetime(2026, 6, 1, 0, 0)
    # A datetime preserves its time-of-day.
    assert isinstance(dt_cell, datetime)
    assert dt_cell == datetime(2026, 6, 1, 13, 30, 15)


def test_custom_sheet_name_with_summary_second_sheet(
    tmp_path: Path,
) -> None:
    """A custom sheet_name plus summary= yields both named sheets."""
    out = tmp_path / "named-summary.xlsx"

    write_xlsx(
        [{"amount": Decimal("12.50"), "currency": "GBP"}],
        out,
        sheet_name="Ledger",
        summary={"count": 1, "total": Decimal("12.50")},
    )

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["Ledger", "Summary"]

    ledger = workbook["Ledger"]
    assert [cell.value for cell in ledger[1]] == ["amount", "currency"]
    assert ledger["A2"].value == pytest.approx(12.5)

    summary = workbook["Summary"]
    assert [cell.value for cell in summary[1]] == ["Key", "Value"]
    assert all(cell.font.bold for cell in summary[1])
    summary_rows = list(summary.iter_rows(min_row=2, values_only=True))
    assert summary_rows[0] == ("count", 1)
    assert summary_rows[1][0] == "total"
    assert summary_rows[1][1] == pytest.approx(12.5)
