# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Example: title the transactions sheet with ``sheet_name``.

The ``sheet_name`` keyword renames the sheet holding the transaction
rows (it defaults to ``"Transactions"``). The workbook is written to a
temporary file, read back with openpyxl to confirm the sheet title, and
the temp file is removed.

Run with ``python examples/05_custom_sheet_name.py``.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from bankstatementparser_writer_xlsx import write_xlsx


def main() -> None:
    """Write to a custom-named sheet and read the title back."""
    records = [
        {"date": "2026-06-01", "description": "Salary", "amount": 3000},
        {"date": "2026-06-03", "description": "Coffee Shop", "amount": -4},
    ]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "ledger.xlsx"
        write_xlsx(records, out, sheet_name="Ledger")

        workbook = load_workbook(out)

        print(f"Wrote {out.name}")
        print(f"Sheets: {workbook.sheetnames}")


if __name__ == "__main__":
    main()
