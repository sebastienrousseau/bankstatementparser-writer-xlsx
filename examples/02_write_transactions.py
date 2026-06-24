# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Example: write a list of ``Transaction`` objects to a workbook.

Transactions are serialised with a stable, documented column order. The
workbook is written to a temporary file, read back with openpyxl to
prove it round-trips, and the temp file is removed.

Run with ``python examples/02_write_transactions.py``.
"""

from __future__ import annotations

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

from bankstatementparser import Transaction
from openpyxl import load_workbook

from bankstatementparser_writer_xlsx import write_xlsx


def main() -> None:
    """Write two transactions, read them back, and print a summary."""
    transactions = [
        Transaction(
            account_id="DE89370400440532013000",
            currency="EUR",
            amount=Decimal("1250.75"),
            booking_date=date(2026, 6, 1),
            description="ACME GmbH invoice 123",
            counterparty="ACME GmbH",
        ),
        Transaction(
            account_id="DE89370400440532013000",
            currency="EUR",
            amount=Decimal("-42.00"),
            booking_date=date(2026, 6, 3),
            description="Coffee Shop",
            counterparty="Coffee Shop",
        ),
    ]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "transactions.xlsx"
        write_xlsx(transactions, out)

        sheet = load_workbook(out)["Transactions"]
        header = [cell.value for cell in sheet[1]]
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))

        print(f"Wrote {out.name}")
        print(f"Header[:3]: {header[:3]}")
        print(f"Rows:       {len(data_rows)}")


if __name__ == "__main__":
    main()
