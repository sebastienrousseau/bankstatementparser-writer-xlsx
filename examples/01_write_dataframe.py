# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Example: write a pandas DataFrame to a single-sheet workbook.

A ``bankstatementparser`` parser returns a DataFrame from ``.parse()``;
this mirrors that workflow with an inline DataFrame so the script runs
with no input files. The workbook is written to a temporary file, read
back with openpyxl to prove it round-trips, and the temp file is removed.

Run with ``python examples/01_write_dataframe.py``.
"""

from __future__ import annotations

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from bankstatementparser_writer_xlsx import write_xlsx


def main() -> None:
    """Write a DataFrame, read it back, and print a summary."""
    frame = pd.DataFrame(
        {
            "date": [date(2026, 6, 1), date(2026, 6, 3)],
            "description": ["Salary", "Coffee Shop"],
            "amount": [Decimal("3000.00"), Decimal("-4.20")],
            "currency": ["EUR", "EUR"],
        }
    )

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "dataframe.xlsx"
        write_xlsx(frame, out)

        sheet = load_workbook(out)["Transactions"]
        header = [cell.value for cell in sheet[1]]
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))

        print(f"Wrote {out.name}")
        print("Sheet:   Transactions")
        print(f"Header:  {header}")
        print(f"Rows:    {len(data_rows)}")


if __name__ == "__main__":
    main()
