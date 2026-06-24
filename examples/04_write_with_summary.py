# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Example: write a workbook with a second ``Summary`` sheet.

Passing ``summary=`` a mapping (e.g. a parser's ``get_summary()``
result) adds a second sheet titled ``Summary`` with a bold
``Key``/``Value`` header. The workbook is written to a temporary file,
read back with openpyxl, and the temp file is removed.

Run with ``python examples/04_write_with_summary.py``.
"""

from __future__ import annotations

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from bankstatementparser_writer_xlsx import write_xlsx


def main() -> None:
    """Write Transactions + Summary sheets and read both back."""
    frame = pd.DataFrame(
        {
            "date": [date(2026, 6, 1), date(2026, 6, 3)],
            "description": ["Salary", "Coffee Shop"],
            "amount": [Decimal("3000.00"), Decimal("-4.20")],
            "currency": ["EUR", "EUR"],
        }
    )
    summary = {
        "account_id": "DE89370400440532013000",
        "transaction_count": len(frame),
        "total_amount": Decimal("2995.80"),
        "currency": "EUR",
    }

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "summary.xlsx"
        write_xlsx(frame, out, summary=summary)

        workbook = load_workbook(out)
        summary_rows = list(
            workbook["Summary"].iter_rows(min_row=2, values_only=True)
        )

        print(f"Wrote {out.name}")
        print(f"Sheets:       {workbook.sheetnames}")
        print(f"Summary rows: {len(summary_rows)}")


if __name__ == "__main__":
    main()
