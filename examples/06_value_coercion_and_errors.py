# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Example: value coercion edge cases and the explicit error paths.

This script demonstrates the parts of ``write_xlsx`` that the other
examples only touch implicitly:

* ``Decimal`` is written as ``float`` and ``date``/``datetime`` as native
  Excel date cells;
* ``None`` and ``float('nan')`` both serialise to a blank cell, and a
  missing ``list[dict]`` key is blank too;
* an over-long string is stored verbatim but its *column width* is capped
  at 60 characters so it never runs off-screen;
* an unsupported top-level type raises ``TypeError`` and an unsupported
  sequence item raises ``ValueError``.

The workbook is written to a temporary file, read back with openpyxl, and
the temp file is removed.

Run with ``python examples/06_value_coercion_and_errors.py``.
"""

from __future__ import annotations

import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from bankstatementparser_writer_xlsx import write_xlsx
from bankstatementparser_writer_xlsx.writer import _MAX_COLUMN_WIDTH


def main() -> None:
    """Show coercion of rich types, blanks, the width cap, and errors."""
    records = [
        {
            "amount": Decimal("1250.75"),  # Decimal -> float
            "booking_date": date(2026, 6, 1),  # native date cell
            "posted_at": datetime(2026, 6, 1, 9, 30),  # native datetime
            "memo": None,  # None -> blank
            "note": "A" * 200,  # over-long -> width capped
        },
        {
            "amount": float("nan"),  # NaN -> blank
            "booking_date": date(2026, 6, 3),
            # "posted_at" missing -> blank
            "memo": "Coffee Shop",
        },
    ]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "coercion.xlsx"
        write_xlsx(records, out)

        sheet = load_workbook(out)["Transactions"]
        header = [cell.value for cell in sheet[1]]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        amount = rows[0][header.index("amount")]
        booking = rows[0][header.index("booking_date")]
        note_width = sheet.column_dimensions[
            get_column_letter(header.index("note") + 1)
        ].width

        assert isinstance(amount, float)  # Decimal coerced to float
        assert isinstance(booking, datetime)  # date read back as datetime
        assert rows[0][header.index("memo")] is None  # None -> blank
        assert rows[1][header.index("amount")] is None  # NaN -> blank
        assert rows[1][header.index("posted_at")] is None  # missing key
        assert note_width == _MAX_COLUMN_WIDTH  # width capped at 60

    # The error paths reject bad input before any file is written.
    type_error = value_error = None
    try:
        write_xlsx("not-a-table", "unused.xlsx")
    except TypeError as exc:
        type_error = exc
    try:
        write_xlsx([42], "unused.xlsx")
    except ValueError as exc:
        value_error = exc

    assert type_error is not None  # bad top-level type
    assert value_error is not None  # bad sequence item

    print(f"Wrote {out.name}")
    print(f"Header:           {header}")
    print(f"Decimal->float:   {amount}")
    print(f"note column width: {note_width} (cap {_MAX_COLUMN_WIDTH})")
    print(f"TypeError:         {type_error}")
    print(f"ValueError:        {value_error}")


if __name__ == "__main__":
    main()
