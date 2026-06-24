# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Example: write a list of plain ``dict`` records to a workbook.

Dict records use the union of their keys in first-seen order, so a key
that only appears in a later record still becomes a column (earlier rows
get a blank cell). The workbook is written to a temporary file, read
back with openpyxl, and the temp file is removed.

Run with ``python examples/03_write_dicts.py``.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from bankstatementparser_writer_xlsx import write_xlsx


def main() -> None:
    """Write dict records, read them back, and print a summary."""
    records = [
        {"date": "2026-06-01", "amount": 3000, "memo": "Salary"},
        {"date": "2026-06-03", "amount": -4, "currency": "EUR"},
    ]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "dicts.xlsx"
        write_xlsx(records, out)

        sheet = load_workbook(out)["Transactions"]
        header = [cell.value for cell in sheet[1]]
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))

        print(f"Wrote {out.name}")
        print(f"Header: {header}")
        print(f"Rows:   {len(data_rows)}")


if __name__ == "__main__":
    main()
