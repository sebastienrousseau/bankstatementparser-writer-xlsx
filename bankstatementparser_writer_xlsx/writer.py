# Copyright (C) 2023-2026 Sebastien Rousseau.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
# implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Excel ``.xlsx`` writer for parsed bank-statement data.

This module turns the output of a
[`bankstatementparser`](https://github.com/sebastienrousseau/bankstatementparser)
parser into a polished Excel workbook via
[openpyxl](https://openpyxl.readthedocs.io/).

The single public entry point is :func:`write_xlsx`, which accepts any
of three input shapes and normalises them to a flat, rectangular table:

* a :class:`pandas.DataFrame` (as returned by a parser's ``.parse()``),
* a list of :class:`bankstatementparser.Transaction` objects, or
* a list of plain ``dict`` row records.

The resulting workbook has a bold header row followed by one row per
record on the ``Transactions`` sheet, with column widths auto-sized for
readability. An optional ``summary`` mapping (e.g. the output of a
parser's ``get_summary()``) is written to a second ``Summary`` sheet as
key/value rows.

Value coercion
--------------
Cells must hold spreadsheet-friendly values, so the writer coerces the
rich Python types the parser emits:

* :class:`decimal.Decimal` is written as a ``float`` (Excel has no
  decimal type; floats sort and aggregate natively in spreadsheets).
* :class:`datetime.date` and :class:`datetime.datetime` are written
  unchanged — openpyxl serialises them as native Excel date cells.
* Every other value is written as-is when openpyxl accepts it
  (``str``, ``int``, ``float``, ``bool``, ``None``) and otherwise
  stringified via ``str()``.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

__all__ = ["write_xlsx"]

# Stable column order for ``Transaction`` rows. Mirrors the field order
# of ``bankstatementparser.Transaction`` so two runs over the same data
# always produce the same workbook layout.
_TRANSACTION_COLUMNS: tuple[str, ...] = (
    "account_id",
    "currency",
    "amount",
    "booking_date",
    "value_date",
    "description",
    "normalized_description",
    "reference",
    "transaction_id",
    "counterparty",
    "source",
    "source_index",
    "source_method",
    "confidence",
    "category",
    "transaction_hash",
)

# Hard upper bound on auto-sized column width, in characters. Wide free
# text (descriptions) would otherwise push columns off-screen.
_MAX_COLUMN_WIDTH = 60


def write_xlsx(
    data: Any,
    path: str | Path,
    *,
    sheet_name: str = "Transactions",
    summary: Mapping[str, Any] | None = None,
) -> Path:
    """Write parsed bank-statement ``data`` to a ``.xlsx`` workbook.

    Args:
        data: The records to serialise. One of:

            * a :class:`pandas.DataFrame` (as returned by a
              ``bankstatementparser`` parser's ``.parse()`` method) —
              its columns define the header order;
            * a list of :class:`bankstatementparser.Transaction`
              objects — serialised with a stable column order;
            * a list of plain ``dict`` records — the header is the
              union of keys in first-seen order.

            An empty list (or empty DataFrame) is accepted and writes a
            header-only sheet (header omitted entirely when no columns
            can be determined).
        path: The output ``.xlsx`` file path. The parent directory must
            already exist; an existing file is overwritten.
        sheet_name: The title of the sheet holding the transaction rows.
            Defaults to ``"Transactions"``.
        summary: Optional mapping (e.g. a parser's ``get_summary()``
            result) written to a second ``"Summary"`` sheet as one
            ``key``/``value`` row per item, under a bold ``Key``/``Value``
            header.

    Returns:
        The :class:`~pathlib.Path` of the written workbook.

    Raises:
        TypeError: If ``data`` is neither a DataFrame, nor a list/tuple
            of ``Transaction`` objects, nor a list/tuple of ``dict``
            records (and is not empty).
        ValueError: If ``data`` is a non-empty sequence whose items mix
            ``dict`` and non-``dict`` types, or contain an unsupported
            item type.
    """
    columns, rows = _normalise(data)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    _write_table(sheet, columns, rows)

    if summary is not None:
        summary_sheet = workbook.create_sheet("Summary")
        _write_summary(summary_sheet, summary)

    output = Path(path)
    workbook.save(output)
    return output


def _normalise(data: Any) -> tuple[list[str], list[list[Any]]]:
    """Normalise any supported ``data`` shape to columns and rows.

    Args:
        data: The input passed to :func:`write_xlsx`.

    Returns:
        A ``(columns, rows)`` pair where ``columns`` is the ordered
        header and ``rows`` is a list of cell-value lists aligned to it.

    Raises:
        TypeError: If ``data`` is of an unsupported top-level type.
        ValueError: If a sequence mixes ``dict`` and non-``dict`` items
            or contains an unsupported item type.
    """
    if isinstance(data, pd.DataFrame):
        return _normalise_dataframe(data)
    if isinstance(data, list | tuple):
        return _normalise_sequence(data)
    raise TypeError(
        "write_xlsx() expects a pandas DataFrame, a list of "
        "Transaction objects, or a list of dict records; got "
        f"{type(data).__name__}"
    )


def _normalise_dataframe(frame: Any) -> tuple[list[str], list[list[Any]]]:
    """Normalise a :class:`pandas.DataFrame` to columns and rows.

    Args:
        frame: The DataFrame to convert. Its column order is preserved.

    Returns:
        A ``(columns, rows)`` pair with one row per DataFrame record.
    """
    columns = [str(column) for column in frame.columns]
    rows = [
        [_coerce(value) for value in record]
        for record in frame.itertuples(index=False, name=None)
    ]
    return columns, rows


def _normalise_sequence(
    records: Sequence[Any],
) -> tuple[list[str], list[list[Any]]]:
    """Normalise a list/tuple of Transactions or dicts to a table.

    Args:
        records: A sequence of ``Transaction`` objects or ``dict``
            records. May be empty.

    Returns:
        A ``(columns, rows)`` pair. An empty sequence yields empty
        columns and rows.

    Raises:
        ValueError: If the items mix ``dict`` and non-``dict`` types or
            an item is neither a ``dict`` nor a ``model_dump``-capable
            object.
    """
    if not records:
        return [], []

    dict_records = [_to_record(record) for record in records]
    columns: list[str] = []
    seen: set[str] = set()
    for record in dict_records:
        for key in record:
            if key not in seen:
                seen.add(key)
                columns.append(key)

    rows = [
        [_coerce(record.get(column)) for column in columns]
        for record in dict_records
    ]
    return columns, rows


def _to_record(record: Any) -> dict[str, Any]:
    """Convert a single sequence item to an ordered ``dict`` record.

    Args:
        record: A ``dict`` or a ``Transaction``-like object exposing a
            ``model_dump()`` method.

    Returns:
        An ordered ``dict`` of column name to value. ``Transaction``
        objects use the stable column order in
        :data:`_TRANSACTION_COLUMNS`; dicts preserve their own order.

    Raises:
        ValueError: If ``record`` is neither a ``dict`` nor exposes a
            callable ``model_dump`` attribute.
    """
    if isinstance(record, Mapping):
        return dict(record)
    model_dump = getattr(record, "model_dump", None)
    if callable(model_dump):
        dumped = model_dump()
        ordered = {
            key: dumped[key] for key in _TRANSACTION_COLUMNS if key in dumped
        }
        for key, value in dumped.items():
            if key not in ordered:
                ordered[key] = value
        return ordered
    raise ValueError(
        "Sequence items must be dict records or Transaction objects "
        f"(with a model_dump() method); got {type(record).__name__}"
    )


def _coerce(value: Any) -> Any:
    """Coerce a Python value to a spreadsheet-friendly cell value.

    Args:
        value: A value drawn from a record.

    Returns:
        ``float`` for :class:`decimal.Decimal`; the value unchanged for
        ``date``/``datetime`` and for the scalar types openpyxl accepts
        natively (``str``, ``int``, ``float``, ``bool``, ``None``);
        ``str(value)`` for anything else.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value
    if value is None or isinstance(value, str | int | float):
        return value
    return str(value)


def _write_table(
    sheet: Any,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    """Write the header and data rows, then auto-size the columns.

    Args:
        sheet: The target openpyxl worksheet.
        columns: The ordered header labels. May be empty, in which case
            nothing is written.
        rows: The cell-value rows aligned to ``columns``.
    """
    if not columns:
        return
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(row)
    _autosize(sheet, columns, rows)


def _write_summary(sheet: Any, summary: Mapping[str, Any]) -> None:
    """Write a ``summary`` mapping as key/value rows on ``sheet``.

    Args:
        sheet: The target openpyxl worksheet (the ``Summary`` sheet).
        summary: The mapping to serialise, one row per item under a
            bold ``Key``/``Value`` header.
    """
    sheet.append(["Key", "Value"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    keys = list(summary)
    for key in keys:
        sheet.append([str(key), _coerce(summary[key])])
    _autosize(
        sheet,
        ["Key", "Value"],
        [[str(key), _coerce(summary[key])] for key in keys],
    )


def _autosize(
    sheet: Any,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    """Set each column width to fit its widest cell, within a cap.

    Args:
        sheet: The worksheet whose column dimensions are adjusted.
        columns: The header labels, used as the initial width seed.
        rows: The data rows scanned for their widest cell per column.
    """
    for index, header in enumerate(columns):
        width = len(str(header))
        for row in rows:
            cell = row[index]
            if cell is not None:
                width = max(width, len(str(cell)))
        letter = get_column_letter(index + 1)
        sheet.column_dimensions[letter].width = min(
            width + 2, _MAX_COLUMN_WIDTH
        )
